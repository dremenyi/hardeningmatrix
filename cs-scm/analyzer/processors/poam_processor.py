
import openpyxl
from typing import List, Dict
from collections import defaultdict
from src.analyzer.models import ComplianceScanResult
from src.cli.utils import ORANGE, RESET

# Import the specialized processors
from .rhel_poam_processor import RhelPoamProcessor
from .postgres_poam_processor import PostgresPoamProcessor

# Create a master registry of all available POAM processors.
ALL_POAM_PROCESSORS = {
    "RHEL 8.X": RhelPoamProcessor,
    "PostgreSQL15_CIS1.1.0": PostgresPoamProcessor,
}


def parse_poam(file_path: str, benchmarks_to_run: List[str]) -> Dict[str, List[ComplianceScanResult]]:
    """
    Parses a POAM (.xlsm) file and groups findings for the specified benchmarks,
    excluding items that are pending review.
    """
    active_processors = []
    for benchmark_name in benchmarks_to_run:
        processor = ALL_POAM_PROCESSORS.get(benchmark_name)
        if processor:
            active_processors.append(processor)
        else:
            print(f"{ORANGE}Warning: No processor found for benchmark '{benchmark_name}'. Skipping.{RESET}")
    
    if not active_processors:
        return {}

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
    except FileNotFoundError:
        print(f"{ORANGE}Error: The file at {file_path} was not found.{RESET}")
        exit(1)
    
    if "Configuration Findings" not in workbook.sheetnames:
        raise ValueError("POAM file must contain a 'Configuration Findings' sheet.")
        
    sheet = workbook["Configuration Findings"]
    
    grouped_results = defaultdict(list)
    
    skipped_findings = []
    
    # Define the column indices to check (R=17, V=21, W=22)
    COLUMNS_TO_CHECK = [17, 21, 22] 

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        is_pending = False
        for col_idx in COLUMNS_TO_CHECK:
            if len(row) > col_idx and row[col_idx]:
                # Check if the cell value contains "pending" (case-insensitive)
                # This handles variations like 'Pending', 'pending review', etc.
                if 'pending' in str(row[col_idx]).lower():
                    is_pending = True
                    break # Stop checking other columns for this row
        
        if is_pending:
            # If the row is pending, add its name to our skipped list and move on.
            weakness_name = row[2] if len(row) > 2 else "Unknown Finding"
            skipped_findings.append(weakness_name)
            continue

        # If the row is not pending, proceed with the normal processor dispatch
        for processor in active_processors:
            if processor.can_process(list(row)):
                result = processor.process(list(row))
                if result:
                    grouped_results[processor.benchmark_name].append(result)
                break
            
    # After checking all rows, print a summary of what was skipped.
    if skipped_findings:
        print(f"{ORANGE}INFO: Skipped {len(skipped_findings)} findings due to 'Pending' status in columns R, V, or W.{RESET}")
        # Optionally, you can print the full list of skipped items here if you want more detail
        # for finding in skipped_findings:
        #     print(f"  - Skipped: {finding}")
            
    return grouped_results