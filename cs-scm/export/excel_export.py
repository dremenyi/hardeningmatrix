"""
Smartsheet Compliance Analyzer - Excel Export Module

This module handles the export of compliance comparison results to Excel format,
with specialized formatting and multiple sheets for different data views.
It now supports creating benchmark-specific sheets for multi-benchmark analysis.
"""
import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict

from src.analyzer.models import ComparisonResult
from src.cli.utils import ORANGE, RESET


def export_to_excel(all_comparison_results: Dict[str, ComparisonResult], output_path: str) -> bool:
    """
    Exports comparison results to a structured Excel report, using shortened
    and sanitized sheet names for each benchmark.
    """
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            create_summary_sheet(writer, all_comparison_results)

            # Define a mapping for full benchmark names to short prefixes
            short_name_map = {
                "RHEL 8.X": "RHEL",
                "PostgreSQL15_CIS1.1.0": "PSQLv15"
                # Add future mappings here
            }

            for benchmark_name, comparison in all_comparison_results.items():
                print(f"--- Generating sheets for {benchmark_name} ---")
                

                # Get the short name from our map, or use the benchmark name as a fallback.
                short_prefix = short_name_map.get(benchmark_name, benchmark_name)
                
                # Sanitize the final prefix to remove all invalid Excel characters.
                safe_prefix = re.sub(r'[\\/*?:\[\]]', '_', short_prefix)


                categorized_data = categorize_findings(comparison)
                
                # Use the new safe prefix to create the sheet titles
                create_report_sheet(writer, f"{safe_prefix}_Approved", categorized_data['approved'])
                create_report_sheet(writer, f"{safe_prefix}_Under_Review", categorized_data['review'])
                create_report_sheet(writer, f"{safe_prefix}_Should_Fix", categorized_data['should_fix'])
                create_report_sheet(writer, f"{safe_prefix}_Un-assessed", categorized_data['unassessed'])
                create_report_sheet(writer, f"{safe_prefix}_Missing_From_SCM", categorized_data['missing_from_scm'])
            
            format_workbook(writer)
        
        return True
    
    except Exception as e:
        print(f"{ORANGE}Error exporting to Excel: {str(e)}{RESET}")
        import traceback
        traceback.print_exc()
        return False

def categorize_findings(comparison: ComparisonResult) -> dict:
    """Sorts all findings from a single comparison object into final lists."""
    findings = {
        'approved': [],
        'review': [],
        'should_fix': [],
        'unassessed': [],
        'missing_from_scm': [item.dict() for item in comparison.unmatched_scan_items]
    }
    
    approved_indicators = ['approved', 'a', 'accept', 'compliant', 'pass']
    review_indicators = ['review', 'pending'] # 'draft' will go to un-assessed

    def categorize_item(item, source=''):
        item_dict = item if isinstance(item, dict) else item.dict()
        if source:
            item_dict['Source'] = source
            

        status = str(item_dict.get('deviation_status', '')).lower().strip()
        should_fix_flag = item_dict.get('should_fix', False)

        
        if should_fix_flag:
            findings['should_fix'].append(item_dict)
        elif any(indicator in status for indicator in approved_indicators):
            findings['approved'].append(item_dict)
        elif any(indicator in status for indicator in review_indicators):
            findings['review'].append(item_dict)
        else:
            # Any other status (including "draft" and blank) will now go here.
            findings['unassessed'].append(item_dict)

    # Process all matched items
    for item in comparison.matched_items:
        categorize_item(item, source='Matched')

    # Process Smartsheet-only items (this list is empty in POAM mode)
    for item in comparison.unmatched_smartsheet_items:
        categorize_item(item, source='Smartsheet Only')
            
    return findings

def create_summary_sheet(writer: pd.ExcelWriter, all_comparison_results: Dict[str, ComparisonResult]):
    """Creates a single summary sheet covering all benchmarks."""
    summary_rows = []
    for benchmark_name, comparison in all_comparison_results.items():
        categorized = categorize_findings(comparison)
        summary_rows.append({'Benchmark': benchmark_name, 'Total Findings': comparison.scan_count, 'Matched': comparison.match_count, 'Approved': len(categorized['approved']), 'Under Review': len(categorized['review']), 'Should Fix': len(categorized['should_fix']), 'Un-assessed': len(categorized['unassessed']), 'Missing from SCM': len(categorized['missing_from_scm'])})
    
    df_summary = pd.DataFrame(summary_rows)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)

def create_report_sheet(writer: pd.ExcelWriter, sheet_name: str, items: list):
    """
    A generic function to create a data sheet from a list of items,
    with special column handling for 'Should Fix' sheets.
    """
    print(f"Exporting {len(items)} items to '{sheet_name}' sheet")
    
    
    # Define the standard list of columns for most sheets
    final_columns = [
        'POAM ID','compliance_id', 'finding_description', 'Hostname', 
        'deviation_type', 'deviation_rationale', 'supporting_documents', 
        'additional_context', 'deviation_status'
    ]

    # If this is a "Should Fix" sheet, add the 'srg_solution' column
    if sheet_name.endswith("_Should_Fix"):
        # Insert the 'srg_solution' column right after 'finding_description'
        final_columns.insert(2, 'srg_solution')

    # This dictionary maps the internal keys to the desired display names for the header
    column_rename_map = {
        'compliance_id': 'Compliance ID',
        'finding_description': 'Finding Description',
        'srg_solution': 'SRG Solution', # Added for Should Fix sheets
        'deviation_type': 'Deviation Type',
        'deviation_rationale': 'Deviation Rationale',
        'supporting_documents': 'Supporting Documents',
        'additional_context': 'Additional Context',
        'deviation_status': 'Deviation Status'
    }

    if not items:
        # Create an empty DataFrame with the correctly formatted final column names
        df = pd.DataFrame(columns=[column_rename_map.get(key, key) for key in final_columns])
    else:
        df = pd.DataFrame(items)
        # Ensure all necessary columns exist, filling missing ones with empty values
        for key in final_columns:
            if key not in df.columns:
                df[key] = None
        
        # Select and order the columns using the internal keys
        df = df[final_columns]
        # Rename the columns to the desired display format
        df = df.rename(columns=column_rename_map)
    
    df.to_excel(writer, sheet_name=sheet_name, index=False)

def format_workbook(writer: pd.ExcelWriter):
    """Applies all the visual formatting to the workbook."""
    workbook = writer.book
    header_fill = PatternFill(start_color='E45E27', end_color='E45E27', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet in workbook.worksheets:
        # Auto-adjust column widths and apply wrapping
        for col_idx, col in enumerate(sheet.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                if cell.row > 1: cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) if max_length > 0 else 20
            sheet.column_dimensions[col_letter].width = min(adjusted_width, 60)

        # Format header
        if sheet.max_row > 0:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.auto_filter.ref = sheet.dimensions
        
        # Apply tab colors based on prefix
        if sheet.title.endswith("_Should_Fix"):
            sheet.sheet_properties.tabColor = 'FF0000' # Red
        elif sheet.title.endswith("_Approved"):
            sheet.sheet_properties.tabColor = 'CCFFCC' # Green
        elif sheet.title.endswith("_Under_Review"):
            sheet.sheet_properties.tabColor = 'FFFFCC' # Yellow

    if 'Summary' in workbook.sheetnames:
        workbook.active = workbook['Summary']