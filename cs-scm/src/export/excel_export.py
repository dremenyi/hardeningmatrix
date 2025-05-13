"""
Smartsheet Compliance Analyzer - Excel Export Module

This module handles the export of compliance comparison results to Excel format,
with specialized formatting and multiple sheets for different data views.

Key Features:
- Multiple sheets for different data views (matched, unmatched, etc.)
- Specialized "Should Fix" sheet for items requiring attention
- Conditional formatting for status and severity
- Data filtering and organization
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
from datetime import datetime
import os

from src.analyzer.models import ComparisonResult, ComplianceItem, ComplianceScanResult
from src.cli.utils import ORANGE, RESET

def export_to_excel(comparison: ComparisonResult, output_path: str) -> bool:
    """
    Export comparison results to a structured Excel report.
    
    Args:
        comparison: ComparisonResult object containing all comparison data
        output_path: Path to save the Excel file
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    try:
        # Create a Pandas Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Add summary sheet
            create_summary_sheet(writer, comparison)
            
            # Add approved items sheet
            create_approved_items_sheet(writer, comparison)
            
            # Add unapproved items sheet
            create_unapproved_items_sheet(writer, comparison)
            
            # Add unmatched scan items sheet
            create_unmatched_scan_sheet(writer, comparison)
            
            # Add dedicated "Should Fix" sheet for items marked should_fix=True
            create_should_fix_sheet(writer, comparison)
            
            # Set column widths and apply formatting
            format_workbook(writer)
        
        return True
    
    except Exception as e:
        print(f"{ORANGE}Error exporting to Excel: {str(e)}{RESET}")
        import traceback
        traceback.print_exc()
        return False


def create_summary_sheet(writer: pd.ExcelWriter, comparison: ComparisonResult):
    """Create a summary sheet with overall statistics."""
    # Count approved and under review items
    approved_matched_count = sum(
        1 for item in comparison.matched_items 
        if not item.get('Should Fix', False) and 
        any(indicator in str(item.get('Deviation Status', '')).lower() 
            for indicator in ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed'])
    )
    
    approved_smartsheet_count = sum(
        1 for item in comparison.unmatched_smartsheet_items
        if not item.should_fix and item.deviation_status and
        any(indicator in str(item.deviation_status).lower() 
            for indicator in ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed'])
    )
    
    under_review_matched_count = sum(
        1 for item in comparison.matched_items
        if not item.get('Should Fix', False) and not
        any(indicator in str(item.get('Deviation Status', '')).lower() 
            for indicator in ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed'])
    )
    
    under_review_smartsheet_count = sum(
        1 for item in comparison.unmatched_smartsheet_items
        if not item.should_fix and (not item.deviation_status or not
        any(indicator in str(item.deviation_status).lower() 
            for indicator in ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed']))
    )
    
    # Create summary data for the first sheet
    summary_data = {
        'Metric': [
            'Scan Date', 
            'Report Generated',
            'Total Scan Items',
            'Total Smartsheet Items',
            'Matched Items',
            'Missing From SCM Template',
            'Approved Items',
            'Items Under Review',
            'Items Requiring Fixes'
        ],
        'Value': [
            comparison.comparison_date.strftime("%Y-%m-%d"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            comparison.scan_count,
            comparison.smartsheet_count,
            comparison.match_count,
            len(comparison.unmatched_scan_items),
            approved_matched_count + approved_smartsheet_count,
            under_review_matched_count + under_review_smartsheet_count,
            sum(1 for item in comparison.matched_items if item.get('Should Fix', False)) +
            sum(1 for item in comparison.unmatched_smartsheet_items if item.should_fix)
        ]
    }
    
    # Convert to DataFrame and write to Excel
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)


def create_approved_items_sheet(writer: pd.ExcelWriter, comparison: ComparisonResult):
    """Create a sheet for approved items that don't need fixing."""
    # Filter for items that are approved and don't need fixing
    approved_items = []
    
    for item in comparison.matched_items:
        # Skip items marked for fixing
        if item.get('Should Fix', False) is True:
            continue
            
        # Check if the item is approved
        deviation_status = str(item.get('Deviation Status', '')).lower()
        is_approved = any(indicator in deviation_status for indicator in 
                        ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed'])
        
        if is_approved:
            approved_items.append(item)
    
    # Add any approved items from unmatched Smartsheet items
    for item in comparison.unmatched_smartsheet_items:
        # Skip items marked for fixing
        if item.should_fix is True:
            continue
            
        # Check if the item is approved
        if item.deviation_status:
            deviation_status = str(item.deviation_status).lower()
            is_approved = any(indicator in deviation_status for indicator in 
                            ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed'])
            
            if is_approved:
                # Convert to dictionary format for DataFrame
                approved_items.append({
                    'Compliance ID': item.compliance_id,
                    'Finding Description': item.finding_description,
                    'SRG Solution': item.srg_solution,
                    'Deviation Type': item.deviation_type,
                    'Deviation Rationale': item.deviation_rationale,
                    'Supporting Documents': item.supporting_documents,
                    'Deviation Status': item.deviation_status,
                    'Source': 'Smartsheet Only',
                    'Should Fix': 'No'
                })
    
    if not approved_items:
        # Create an empty DataFrame with column headers
        df_approved = pd.DataFrame(columns=[
            'Compliance ID', 'Status', 'Severity', 'Hostname',
            'Finding Description', 'SRG Solution', 'Deviation Type',
            'Deviation Rationale', 'Supporting Documents', 'Deviation Status', 'Source'
        ])
    else:
        df_approved = pd.DataFrame(approved_items)
    
    print(f"Exporting {len(approved_items)} items to 'Approved Items' sheet")
    df_approved.to_excel(writer, sheet_name='Approved Items', index=False)


def create_unapproved_items_sheet(writer: pd.ExcelWriter, comparison: ComparisonResult):
    """Create a sheet for unapproved/pending items that don't need fixing."""
    # Filter for items that are not approved and don't need fixing
    unapproved_items = []
    
    for item in comparison.matched_items:
        # Skip items marked for fixing
        if item.get('Should Fix', False) is True:
            continue
            
        # Check if the item is not approved
        deviation_status = str(item.get('Deviation Status', '')).lower()
        approved_indicators = ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed']
        is_approved = any(indicator in deviation_status for indicator in approved_indicators)
        
        if not is_approved:
            unapproved_items.append(item)
    
    # Add any unapproved items from unmatched Smartsheet items
    for item in comparison.unmatched_smartsheet_items:
        # Skip items marked for fixing
        if item.should_fix is True:
            continue
            
        # Check if the item is not approved
        if item.deviation_status:
            deviation_status = str(item.deviation_status).lower()
            approved_indicators = ['approved', 'a', 'accept', 'acceptable', 'compliant', 'comp', 'pass', 'passed']
            is_approved = any(indicator in deviation_status for indicator in approved_indicators)
            
            if not is_approved:
                # Convert to dictionary format for DataFrame
                unapproved_items.append({
                    'Compliance ID': item.compliance_id,
                    'Finding Description': item.finding_description,
                    'SRG Solution': item.srg_solution,
                    'Deviation Type': item.deviation_type,
                    'Deviation Rationale': item.deviation_rationale,
                    'Supporting Documents': item.supporting_documents,
                    'Deviation Status': item.deviation_status or 'Pending Review',
                    'Source': 'Smartsheet Only',
                    'Should Fix': 'No'
                })
    
    if not unapproved_items:
        # Create an empty DataFrame with column headers
        df_unapproved = pd.DataFrame(columns=[
            'Compliance ID', 'Status', 'Severity', 'Hostname',
            'Finding Description', 'SRG Solution', 'Deviation Type',
            'Deviation Rationale', 'Supporting Documents', 'Deviation Status', 'Source',
            'Should Fix'
        ])
    else:
        df_unapproved = pd.DataFrame(unapproved_items)
    
    print(f"Exporting {len(unapproved_items)} items to 'Under Review' sheet")
    df_unapproved.to_excel(writer, sheet_name='Under Review', index=False)


def create_unmatched_scan_sheet(writer: pd.ExcelWriter, comparison: ComparisonResult):
    """Create a sheet for unmatched scan items."""
    # Convert scan items to dictionaries
    scan_items = [
        {
            'Compliance ID': item.compliance_id,
            'Status': item.status,
            'Severity': item.severity,
            'Hostname': item.hostname,
            'Description': item.description,
            'Details': item.details,
            **item.additional_fields
        }
        for item in comparison.unmatched_scan_items
    ]
    
    if not scan_items:
        # Create an empty DataFrame with column headers
        df_scan = pd.DataFrame(columns=[
            'Compliance ID', 'Status', 'Severity', 'Hostname',
            'Description', 'Details'
        ])
    else:
        df_scan = pd.DataFrame(scan_items)
    
    df_scan.to_excel(writer, sheet_name='Missing From SCM Template', index=False)


# This function has been replaced by create_approved_items_sheet and create_unapproved_items_sheet
# The unmatched Smartsheet items are now included in those sheets based on their approval status


def create_should_fix_sheet(writer: pd.ExcelWriter, comparison: ComparisonResult):
    """Create a dedicated sheet for items that should be fixed."""
    # Get matched items that should be fixed
    matched_fix_items = [
        {
            'Source': 'Matched',
            'Compliance ID': item.get('Compliance ID'),
            'Finding Description': item.get('Finding Description'),
            'SRG Solution': item.get('SRG Solution'),
            'Deviation Type': item.get('Deviation Type'), 
            'Deviation Rationale': item.get('Deviation Rationale'),
            'Status': item.get('Status'),
            'Severity': item.get('Severity'),
            'Hostname': item.get('Hostname'),
            'Comments': item.get('Comments', ''),
            'Should Fix': 'Yes'
        }
        for item in comparison.matched_items 
        if item.get('Should Fix', False) is True
    ]
    
    # Get unmatched Smartsheet items that should be fixed
    unmatched_fix_items = [
        {
            'Source': 'Unmatched Smartsheet',
            'Compliance ID': item.compliance_id,
            'Finding Description': item.finding_description,
            'SRG Solution': item.srg_solution,
            'Deviation Type': item.deviation_type,
            'Deviation Rationale': item.deviation_rationale,
            'Status': 'N/A',
            'Severity': 'N/A',
            'Hostname': 'N/A',
            'Comments': item.comments or '',
            'Should Fix': 'Yes'
        }
        for item in comparison.unmatched_smartsheet_items 
        if item.should_fix is True
    ]
    
    # Combine both lists
    all_fix_items = matched_fix_items + unmatched_fix_items
    
    if not all_fix_items:
        # Create an empty DataFrame with column headers
        df_fix = pd.DataFrame(columns=[
            'Source', 'Compliance ID', 'Finding Description', 'SRG Solution',
            'Deviation Type', 'Deviation Rationale', 'Status', 'Severity',
            'Hostname', 'Comments', 'Should Fix'
        ])
    else:
        df_fix = pd.DataFrame(all_fix_items)
    
    print(f"Exporting {len(all_fix_items)} items to 'Should Fix Items' sheet")
    df_fix.to_excel(writer, sheet_name='Should Fix Items', index=False)


def format_workbook(writer: pd.ExcelWriter):
    """Apply formatting to the workbook."""
    workbook = writer.book
    
    # Define styles
    header_fill = PatternFill(start_color='E45E27', end_color='E45E27', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define alternating row colors
    alt_row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alt_row_fill_grey = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light grey
    
    # Status color mapping
    status_colors = {
        'approved': 'CCFFCC',  # Light green
        'pass': 'CCFFCC',
        'passed': 'CCFFCC',
        'compliant': 'CCFFCC',
        'a': 'CCFFCC',
        'pending': 'FFFFCC',  # Light yellow
        'review': 'FFFFCC',
        'pending review': 'FFFFCC',
        'rejected': 'FFCCCC',  # Light red
        'fail': 'FFCCCC',
        'failed': 'FFCCCC',
        'non-compliant': 'FFCCCC'
    }
    
    # Severity color mapping
    severity_colors = {
        'critical': 'FF9999',  # Red
        'high': 'FFCC99',      # Orange
        'medium': 'FFFFCC',    # Yellow
        'low': 'CCFFCC',       # Green
        'informational': 'CCCCFF'  # Blue
    }
    
    # Format each worksheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Set column widths
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            # Adjust width with a little padding
            adjusted_width = (max_length + 2) if max_length > 0 else 15
            worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)
        
        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add auto-filter
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Get header indices for colored columns
        status_col_idx = None
        severity_col_idx = None
        deviation_status_col_idx = None
        
        for idx, cell in enumerate(worksheet[1], 1):
            if cell.value == 'Status':
                status_col_idx = idx
            elif cell.value == 'Severity':
                severity_col_idx = idx
            elif cell.value == 'Deviation Status':
                deviation_status_col_idx = idx
        
        # Format rows with alternating colors and apply conditional formatting
        # Start from row 2 (after header) or row 3 if there's a banner (in special sheets)
        start_row = 2
        if sheet_name in ['Should Fix Items', 'Approved Items', 'Under Review']:
            start_row = 3
            
        for row_idx in range(start_row, worksheet.max_row + 1):
            # Apply alternating row color
            row_fill = alt_row_fill_grey if row_idx % 2 == 0 else alt_row_fill_white
            
            for cell in worksheet[row_idx]:
                # Add border to all cells
                cell.border = border
                
                # Set alternating row background (will be overridden by status colors if applicable)
                cell.fill = row_fill
                
                # Set wrapping for text cells
                if isinstance(cell.value, str) and len(cell.value) > 40:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Only override row color for specific status cells that need highlighting
                
                # Apply status coloring
                if status_col_idx and cell.column == status_col_idx and cell.value:
                    status_val = str(cell.value).lower()
                    for status_key, color in status_colors.items():
                        if status_key in status_val:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            break
                
                # Apply severity coloring
                if severity_col_idx and cell.column == severity_col_idx and cell.value:
                    severity_val = str(cell.value).lower()
                    for severity_key, color in severity_colors.items():
                        if severity_key in severity_val:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            break
                
                # Apply deviation status coloring
                if deviation_status_col_idx and cell.column == deviation_status_col_idx and cell.value:
                    deviation_val = str(cell.value).lower()
                    for status_key, color in status_colors.items():
                        if status_key in deviation_val:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            break
        
        # Special sheet formatting
        if sheet_name == 'Should Fix Items':
            # Set tab color to red
            worksheet.sheet_properties.tabColor = 'FF0000'
            
            # Add a note at the top of the sheet
            worksheet.insert_rows(1)
            top_note = worksheet.cell(row=1, column=1)
            top_note.value = "ATTENTION: All items in this sheet require action"
            top_note.font = Font(bold=True, color='FF0000', size=12)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(worksheet[2]))
            top_note.alignment = Alignment(horizontal='center')
            
            # Bold the "Items Requiring Fixes" text in summary if it exists
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                for row in summary_sheet.iter_rows():
                    for cell in row:
                        if cell.value == 'Items Requiring Fixes':
                            cell.font = Font(bold=True, color='FF0000')
                            # Make the value cell red too
                            value_cell = summary_sheet.cell(row=cell.row, column=cell.column + 1)
                            value_cell.font = Font(bold=True, color='FF0000')
                            
                            # Add hyperlink to the Should Fix sheet
                            value_cell.hyperlink = f"#{sheet_name}!A1"
                            value_cell.style = "Hyperlink"
        
        elif sheet_name == 'Approved Items':
            # Set tab color to green
            worksheet.sheet_properties.tabColor = 'CCFFCC'
            
            # Add a note at the top of the sheet
            worksheet.insert_rows(1)
            top_note = worksheet.cell(row=1, column=1)
            top_note.value = "These items have approved deviations - no action required"
            top_note.font = Font(bold=True, color='006600', size=12)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(worksheet[2]))
            top_note.alignment = Alignment(horizontal='center')
            
            # Link from summary
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                for row in summary_sheet.iter_rows():
                    for cell in row:
                        if cell.value == 'Approved Items':
                            # Add hyperlink to the Approved Items sheet
                            value_cell = summary_sheet.cell(row=cell.row, column=cell.column + 1)
                            value_cell.hyperlink = f"#{sheet_name}!A1"
                            value_cell.style = "Hyperlink"
        
        elif sheet_name == 'Under Review':
            # Set tab color to yellow
            worksheet.sheet_properties.tabColor = 'FFFFCC'
            
            # Add a note at the top of the sheet
            worksheet.insert_rows(1)
            top_note = worksheet.cell(row=1, column=1)
            top_note.value = "These items are pending review or need discussion"
            top_note.font = Font(bold=True, color='CC9900', size=12)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(worksheet[2]))
            top_note.alignment = Alignment(horizontal='center')
            
            # Link from summary
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                for row in summary_sheet.iter_rows():
                    for cell in row:
                        if cell.value == 'Items Under Review':
                            # Add hyperlink to the Under Review sheet
                            value_cell = summary_sheet.cell(row=cell.row, column=cell.column + 1)
                            value_cell.hyperlink = f"#{sheet_name}!A1"
                            value_cell.style = "Hyperlink"
    
    # Make Summary the active sheet when opening
    if 'Summary' in workbook.sheetnames:
        workbook.active = workbook.worksheets[workbook.sheetnames.index('Summary')]